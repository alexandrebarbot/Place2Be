#!/usr/local/bin/python3

import requests
import os.path, sys

from openpyxl import load_workbook
from geopy.geocoders import Nominatim


from _database import connection, insert_value, select_value

geolocator = Nominatim()

"""

	DATA USED 		: https://data.public.lu/fr/datasets/population-par-commune-et-code-lau2-depuis-2000/#_
	URL 			: https://download.data.public.lu/resources/population-par-commune-et-code-lau2-depuis-2000/20170511-141118/popcom2000-2017_LAU2.xlsx
	DESCRIPTION 	: Chiffres de population par commune et code LAU2 depuis 2000

"""

url = "https://download.data.public.lu/resources/population-par-commune-et-code-lau2-depuis-2000/20170511-141118/popcom2000-2017_LAU2.xlsx"
target_path = "tmp/data_people_per_town.xlsx"



def _download():
	''' Download xlsx file from data.public.lu '''
	global target_path, url

	response = requests.get(url, stream=True)
	handle = open(target_path, "wb")
	for chunk in response.iter_content(chunk_size=512):
	    if chunk:  # filter out keep-alive new chunks
	        handle.write(chunk)

	if os.path.isfile(target_path):
		return True
	else:
		sys.exit("File not found")


def get_city_location(city):
	''' Return the lattitude and longitude of the current city '''
	global geolocator

	location = geolocator.geocode(city)

	try:
		return location.latitude, location.longitude
	except:
		print("City unknow, manual insert", city)
		#sys.exit()

def parse_xlsx():
	workbook = load_workbook(target_path)
	first_sheet = workbook.get_sheet_names()[0]
	worksheet = workbook.get_sheet_by_name(first_sheet)

	for row in worksheet.iter_rows("C6:U110"):
		city = row[0].value

		sql = "SELECT id from cities WHERE name = '%s'" % city
		pcity = select_value(sql)
		# Check if the city already exist
		if not pcity:
			try:
				lat, lon = get_city_location(row[0].value)
			except:
				continue
			print("CITY", city, "Lat", lat, 'Lon', lon)

			# Insert City is does not exist
			sql = "INSERT INTO cities VALUES (NULL, '%s', %s, %s)" % (city, lat, lon)
			city_id = insert_value(sql)
		else:
			city_id = pcity['id']

		if not city_id:
			continue

		# Insert City Demography
		year = 2000
		for cell in row[1:]:
			if cell.value > 10:
				sql = "INSERT INTO populations VALUES (NULL, %s, %s, %s) ON DUPLICATE KEY UPDATE value = %s;" % (city_id, cell.value, year, cell.value)
				insert_value(sql)
				print(sql)
			else:
				continue
			year += 1

if __name__ == '__main__':
	parse_xlsx()







